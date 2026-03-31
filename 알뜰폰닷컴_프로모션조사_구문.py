# ──────────────────────────────────────────────────────────────────────────────
# 1. 임포트 영역 (import ...)
# ──────────────────────────────────────────────────────────────────────────────

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import re
import pandas as pd
import numpy as np
import smtplib
from bs4 import BeautifulSoup
from datetime import datetime, timezone, timedelta
from email.message import EmailMessage
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# ──────────────────────────────────────────────────────────────────────────────
# 2. 설정값 영역 (ID, PW, 선택자 등)
# ──────────────────────────────────────────────────────────────────────────────

EMAIL_PASSWORD = "udvi vntq lfbq tjsq"         # Gmail 앱 비밀번호 (Google 계정 설정에서 생성)
SENDER_EMAIL = "doda427@gmail.com"           # 발신자 이메일 주소
RECEIVER_EMAIL = [ "samlee1018@lguplus.co.kr" ]         # 수신자 이메일 주소
MOYOPLAN_BASE_URL = "https://www.uplusmvno.com/plan/plan-list" # 알뜰폰닷컴 기본 URL 설정
SCROLL_PAUSE_SEC = 1 # 스크롤 대기 시간 설정
PAGE_LOAD_DELAY = 1.5 # 페이지 로딩 대기 시간
VALID_TELECOM_NAMES = ['SKT', 'LG U+', 'KT'] # 유효한 통신사 목록
PLAN_CONTAINER_SELECTOR = "div[class*='tny78c4']"           # 요금제 컨테이너
MVNO_INFO_SELECTOR = "img[class*='h-4']"        # MVNO사업자명 정보
NAME_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3dz _1u9yo3d7 _1u9yo3dr']"        # 요금제명 정보
DATA_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3d11']"               # 데이터 크기 및 속도 정보
DISCOUNT_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3dz']"             # 할인 개월수 및 원가 정보
CALL_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3dz _1u9yo3d7']"              # 통화 제공량 정보
MMS_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3dz _1u9yo3d7']"              # 문자 제공량 정보
TELECOM_INFO_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3dz _1u9yo3d7']"           # 통신사 정보
PRICE_SELECTOR = "span[class*='_1u9yo3d4 _1u9yo3d17']"                   # 프로모션 가격
PAGE_BUTTON_SELECTOR = "a[class*='flex'][class*='items-center'][class*='justify-center'][href*='page=']"                         # 페이지 버튼 (flex + items-center + justify-center + page 파라미터 모두 포함)
KST = timezone(timedelta(hours=9))
SIDE_MARGIN = 5  # 엑셀 열 너비 추가 여유분
ROW_HEIGHT = 15  # 엑셀 행 높이 설정

# ──────────────────────────────────────────────────────────────────────────────
# 3. 유틸리티 함수 영역 (클래스에 속하지 않는 독립적 기능)
# ──────────────────────────────────────────────────────────────────────────────

def setup_driver():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--window-size=1920,1080')
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    return webdriver.Chrome(options=options)

def extract_plan_data(soup):
    results = []
    containers = soup.select("ul.plan-list > li.item")
    
    for container in containers:
        try:
            # 1. MVNO 사업자명 (로고 이미지의 alt 속성)
            mvno_img = container.select_one(".plan-top .logo img")
            mvno_name = mvno_img.get('alt', '기타').strip() if mvno_img else "기타"
            
            # 2. 요금제명
            name_tag = container.select_one(".plan-tit")
            name = name_tag.get_text(strip=True) if name_tag else "정보없음"
            
            # 3. 데이터 및 QoS 정보 파싱
            data_info_tag = container.select_one(".plan-info .data")
            data_info = data_info_tag.get_text(strip=True) if data_info_tag else ""
            
            # 데이터 크기 (숫자 + GB/MB)
            data_match = re.search(r'(\d+\.?\d*)\s*(GB|MB)', data_info, re.I)
            refined_data = data_match.group(0) if data_match else "0GB"
            
            # QoS (속도제한)
            qos_match = re.search(r'\d+\s*[kM]?bps', data_info, re.I)
            qos_val = qos_match.group() if qos_match else ""
            
            # 4. 통화/문자 제공량
            info_list = container.select(".info-list li")
            call, mms = "정보없음", "정보없음"
            for li in info_list:
                txt = li.get_text(strip=True)
                if "통화" in txt: call = txt.replace("통화", "").strip()
                if "문자" in txt: mms = txt.replace("문자", "").strip()

            # 5. 가격 정보 (프로모션가, 원가, 할인개월)
            # 프로모션가 (현재 구매가)
            price_tag = container.select_one(".price-area .p-price b")
            price_raw = "".join(re.findall(r'\d+', price_tag.text)) if price_tag else "0"
            
            # 할인 정보 (개월수 추출)
            discount_tag = container.select_one(".price-area .p-info")
            discount_text = discount_tag.get_text(strip=True) if discount_tag else ""
            discount_months = "평생"
            if "개월" in discount_text:
                month_match = re.search(r'(\d+)개월', discount_text)
                if month_match: discount_months = month_match.group(1)
            
            # 원가 (취소선 가격)
            original_price_tag = container.select_one(".price-area .p-del")
            if original_price_tag:
                original_price = "".join(re.findall(r'\d+', original_price_tag.text))
            else:
                original_price = price_raw # 원가 정보 없으면 프로모션가와 동일하게 처리

            results.append({
                "MVNO사업자명": mvno_name,
                "통신사": "LG U+",
                "요금제명": name,
                "통화": call,
                "문자": mms,
                "데이터": refined_data,
                "QoS": qos_val,
                "원가": int(original_price) if original_price.isdigit() else 0,
                "할인개월": discount_months,
                "프로모션가": int(price_raw) if price_raw.isdigit() else 0
            })
        except: continue
    return results

def send_email_report(excel_filename):
    now = datetime.now(KST)
    today = now.strftime("%m%d")
    mail_date = now.strftime("%Y-%m-%d")
    try:
        msg = EmailMessage()
        msg['Subject'] = f"모요_전체요금제_{today}"
        msg['From'] = SENDER_EMAIL
        msg['To'] = ", ".join(RECEIVER_EMAIL) if isinstance(RECEIVER_EMAIL, list) else RECEIVER_EMAIL
        msg.set_content(f"안녕하세요,\n\n{mail_date} 기준 알뜰폰닷컴 전체 요금제 수집 결과입니다.")

        with open(excel_filename, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=excel_filename
            )

        with smtplib.SMTP("smtp.gmail.com", 587) as connection:
            connection.starttls()
            connection.login(SENDER_EMAIL, EMAIL_PASSWORD)
            connection.send_message(msg)
        return True
    except Exception as e:
        print(f"❌ 이메일 발송 실패: {e}")
        return False

# ──────────────────────────────────────────────────────────────────────────────
# 4. 크롤링 클래스 영역
# ──────────────────────────────────────────────────────────────────────────────

class MoyoplanScraper:
    def __init__(self):
        self.driver = setup_driver()

    def scroll_to_bottom(self):
        try:
            last_height = self.driver.execute_script("return document.body.scrollHeight")
            while True:
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(SCROLL_PAUSE_SEC)
                new_height = self.driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
        except Exception as e:
            print(f"스크롤 중 오류 발생: {e}")

    def scrape_all(self):
        all_data, visited = [], set()
        self.driver.get(MOYOPLAN_BASE_URL)
        time.sleep(PAGE_LOAD_DELAY)

        while True:
            # 1. 기존의 안정적인 스크롤로 바닥까지 내려서 데이터/버튼 활성화
            self.scroll_to_bottom()
            
            # 2. 현재 페이지 데이터 수집
            current_soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            all_data.extend(extract_plan_data(current_soup))
            
            # 3. 페이지 버튼들 찾기 (기존 PAGE_BUTTON_SELECTOR 활용)
            page_buttons = self.driver.find_elements(By.CSS_SELECTOR, PAGE_BUTTON_SELECTOR)
            target_btn = None
            current_num = None
            
            # 방문하지 않은 숫자 버튼 찾기
            for btn in page_buttons:
                txt = btn.text.strip()
                if txt.isdigit():
                    num = int(txt)
                    if num not in visited:
                        target_btn = btn
                        current_num = num
                        break
            
            if target_btn:
                visited.add(current_num)
                # 버튼이 클릭 가능하도록 스크롤 후 클릭
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_btn)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", target_btn)
                print(f"✅ {current_num}페이지 수집 완료 (누적: {len(all_data)}건)")
                time.sleep(PAGE_LOAD_DELAY)
            else:
                # 4. 숫자 버튼이 없으면 '다음(>)' 블록 이동 시도
                print("➔ 현재 블록 완료. 다음 블록(>) 확인 중...")
                try:
                    # 화살표 아이콘(ChevronRight)을 가진 버튼 찾기
                    next_block_xpath = "//*[local-name()='svg' and contains(@data-sentry-component, 'ChevronRightIcon')]/ancestor::button | //*[local-name()='svg' and contains(@data-sentry-component, 'ChevronRightIcon')]/ancestor::a"
                    next_btn = self.driver.find_element(By.XPATH, next_block_xpath)
                    self.driver.execute_script("arguments[0].click();", next_btn)
                    print("▶ 다음 페이지 블록으로 이동합니다.")
                    time.sleep(3) # 블록 이동 후 로딩 대기
                except:
                    print("✨ 모든 데이터 수집이 완료되었습니다.")
                    break

        self.driver.quit()
        df = pd.DataFrame(all_data)
        if not df.empty:
            # extract_plan_data에서 정의한 키값과 100% 일치해야 함
            df = df.drop_duplicates(subset=['요금제명', '통신사', '프로모션가'])
        return df        

# ──────────────────────────────────────────────────────────────────────────────
# 5. 실행 영역 (메인 함수)
# ──────────────────────────────────────────────────────────────────────────────

def main():
    now = datetime.now(KST)
    today = now.strftime("%m%d")
    print(f"🚀 수집 시작 시간: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    try:
        # 1. 수집
        scraper = MoyoplanScraper()
        final_df = scraper.scrape_all()
        
        if final_df.empty:
            print("⚠️ 수집된 데이터가 없습니다.")
            return

        # 2. 데이터 정제 (가격 정렬을 위해 숫자로 변환)
        if '프로모션가' in final_df.columns:
            final_df['프로모션가'] = pd.to_numeric(final_df['프로모션가'], errors='coerce')

        # 3. 저장 및 정렬
        filename = f"moyo_전체요금제_{today}.xlsx"
        final_df.sort_values(by=['MVNO사업자명', '통신사', '프로모션가'], ascending=[True, True, True]).to_excel(
            filename, index=False
        )
        print(f"💾 기본파일 저장 완료: {filename} (총 {len(final_df)}건)")
        
        # --- 🎨 엑셀 디자인 최적화 시작 ---
        try:
            wb = load_workbook(filename)
            ws = wb.active

            custom_border = Border(
                left=Side(style='thin', color='000000'),   # 좌측 실선
                right=Side(style='thin', color='000000'),  # 우측 실선
                bottom=Side(style='thin', color='000000'),    # 상단 실선
            )

            # 1. 모든 행 높이 및 중앙 정렬 설정
            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = ROW_HEIGHT
                for cell in row:
                    cell.border = custom_border
                    cell.font = Font(name='맑은 고딕', size=9)
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 2. 열 너비 자동 조절 로직
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter # A, B, C...
                
                for cell in col:
                    if cell.value:
                        # 한글은 영문보다 자리를 많이 차지하므로 길이를 넉넉히 계산
                        val_str = str(cell.value)
                        current_length = 0
                        for char in val_str:
                            if ord(char) > 128: # 한글 등 멀티바이트 문자
                                current_length += 2
                            else:
                                current_length += 1.2
                        
                        if current_length > max_length:
                            max_length = current_length
                
                # 너무 넓어지는 것 방지 (최대 50)
                adjusted_width = max(10, min(max_length + SIDE_MARGIN, 50))
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            print(f"✨ 엑셀 디자인 최적화 완료 (줄간격/정렬/너비)")

        except Exception as e:
            print(f"⚠️ 엑셀 디자인 적용 중 실제 오류 발생: {e}")

        # 4. 발송
        if send_email_report(filename):
            print("📧 이메일 발송 완료!")
            
    except Exception as e:
        print(f"🚨 메인 로직 오류: {e}")

if __name__ == "__main__":
    main()
